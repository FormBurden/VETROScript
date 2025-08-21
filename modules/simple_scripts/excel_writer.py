# modules/simple_scripts/excel_writer.py

import logging
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
import re
import modules.config
import glob
from collections import Counter

from modules.simple_scripts.slack_loops import find_underground_slack_mismatches, find_slack_dist_mismatches
from modules.simple_scripts.distribution import find_missing_distribution_footage
from modules.simple_scripts.fiber_drop import find_missing_service_location_drops
from modules.simple_scripts.footage_issues import find_overlength_drop_cables
from modules.config import ID_COL
from modules.simple_scripts.nap_rules import find_nap_drop_mismatches, find_nap_id_format_issues, scan_nap_spec_warnings


logger = logging.getLogger(__name__)

def new_workbook():
    wb = Workbook()
    return wb, wb.active


def auto_size(wb):
    for ws in wb.worksheets:
        # ─── skip our manual‐width sheet ────────────────────────────
        if ws.title == 'Power Pole Issues':
            continue
        for col in ws.columns:
            # on Drop Issues, ignore the long description in row 1
            cells = col[1:] if ws.title == 'Drop Issues' else col
            w = max((len(str(c.value)) for c in cells if c.value), default=0) + 2
            ws.column_dimensions[get_column_letter(col[0].column)].width = w    

def drop_empty_issue_sheets(wb):
    """
    Remove any issue-oriented worksheet that ended up empty (no data rows),
    unless SHOW_ALL_SHEETS is True.

    A sheet is considered "issue-oriented" if:
      • Its title contains "Issue" or "Issues", OR
      • It matches one of the common issue sheet names below.

    Data check: any non-empty cell exists at/after row 2 (row 1 is usually headers).
    """
    import modules.config

    if getattr(modules.config, "SHOW_ALL_SHEETS", False):
        return

    COMMON_ISSUE_SHEETS = {
        "Drop Issues",
        "Slack Loop Issues",
        "Conduit Issues",
        "Vault Issues",
        "Power Pole Issues",
        "NID Issues",
        "Footage Issues",
        "Distribution and NAP Walker",
    }

    def _is_issue_sheet(title: str) -> bool:
        t = str(title or "")
        return ("Issue" in t) or (t in COMMON_ISSUE_SHEETS)

    # don’t touch informational sheets
    KEEP_ALWAYS = {"PON Statistics", "GeoJSON Summary"}

    for ws in list(wb.worksheets):
        if ws.title in KEEP_ALWAYS:
            continue
        if not _is_issue_sheet(ws.title):
            continue

        # Any non-empty cell from row 2 down?
        has_data = False
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if any(v not in (None, "") for v in row):
                has_data = True
                break

        if not has_data:
            wb.remove(ws)


def natural_key(s: str):
    """Split a string into text and number chunks for natural ordering."""
    parts = re.split(r'(\d+)', s)
    return [int(p) if p.isdigit() else p.lower() for p in parts]


def write_network_statistics(wb, stats):
    """
    Inserts a ‘PON Statistics’ sheet at the front (index=0) and writes:
      • Left block (A:B): key network metrics + issue counts
        - Metrics (NAPs..Vaults) alphabetized
        - 'T3 Vault' stays fixed after the metrics
        - Issue rows (Fiber-Drop Issues..NAP Issues) alphabetized
      • Right block (D:E): "PON Layers Missing/Present" with existence checks (alphabetized)

    Formatting:
      • Row 1 merged A1:B1 title "Misc Network Info"
      • Row 2 column headers "Metric", "Value"
      • Value column (B) centered
      • Issue rows bolded when count > 0
      • Borders applied by apply_borders(ws)
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    import glob
    import modules.config  # required: use modules.config.DATA_DIR

    # 1) Create the sheet at index 0
    ws = wb.create_sheet(title='PON Statistics', index=0)
    ws.freeze_panes = 'A3'  # freeze title + column header rows

    # 2) Big merged title in row 1 (A1:B1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    header = ws.cell(row=1, column=1, value='Misc Network Info')
    header.alignment = Alignment(horizontal='center')
    header.font = Font(bold=True)

    # 3) Column titles on row 2 (bold + centered)
    cols = ['Metric', 'Value']
    for col_idx, title in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=col_idx, value=title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # ---------- Left block (A:B) ----------
    # Metrics (alphabetize NAPs..Vaults)
    metrics_base = {
        'NAPs': stats.get('nap_count', 0),
        'Service Locations': stats.get('service_location_count', 0),
        'NIDs': stats.get('nid_count', 0),
        'Power Poles': stats.get('power_pole_count', 0),
        'Vaults': stats.get('vault_count_excluding_t3', 0),
    }
    metrics_sorted = sorted(metrics_base.items(), key=lambda kv: kv[0].lower())

    # T3 Vault (stays right after the metrics)
    t3_names = stats.get('t3_names', [])
    if isinstance(t3_names, (list, tuple)):
        t3_joined = ', '.join(t3_names)
    else:
        t3_joined = str(t3_names) if t3_names is not None else ''

    # Issues (alphabetize Fiber-Drop Issues..NAP Issues)
    nap_issue_mismatch = stats.get('nap_mismatch_issues', stats.get('nap_mismatches', 0))
    nap_issues_total = nap_issue_mismatch + stats.get('nap_naming_issues', 0) + stats.get('nap_spec_warnings', 0)
    slack_issues_total = (
        stats.get('slack_dist_issues', 0)
        + stats.get('underground_slack_issues', 0)
        + stats.get('aerial_slack_issues', 0)
        + stats.get('tail_end_slack_issues', 0)
    )
    issues_base = {
        'Conduit Issues': stats.get('conduit_issues', 0),
        'Dist/NAP Walker Issues': stats.get('dist_nap_walker_issues', 0),
        'Fiber-Drop Issues': stats.get('fiber_drop_issues', 0),
        'Footage Issues': stats.get('footage_issues', 0),
        'NID Drop Issues': stats.get('nid_drop_issues', 0),
        'NAP Issues': nap_issues_total,
        'Power Pole Issues': stats.get('power_pole_issues', 0),
        'SL Attributes Issues': stats.get('svc_attr_issues', 0),
        'Slack Loop Issues': slack_issues_total,
        'Vault Issues': stats.get('vault_issues', 0),
    }
    issues_sorted = sorted(issues_base.items(), key=lambda kv: kv[0].lower())

    # Write rows
    start_row = 3
    r = start_row

    # Metrics (alphabetized)
    for label, val in metrics_sorted:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=val)
        r += 1

    # T3 Vault (fixed position)
    ws.cell(row=r, column=1, value='T3 Vault')
    ws.cell(row=r, column=2, value=t3_joined)
    r += 1

    issue_start_idx = r  # first issue row index

    # Issues (alphabetized)
    for label, val in issues_sorted:
        cell_label = ws.cell(row=r, column=1, value=label)
        cell_value = ws.cell(row=r, column=2, value=val)

        # Bold any issue row when count > 0
        if isinstance(val, int) and val > 0:
            cell_label.font = Font(bold=True)
            cell_value.font = Font(bold=True)
        r += 1

    # Center-align the Value column
    for row in ws.iter_rows(min_row=1, min_col=2, max_col=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    # ---------- Right block (D:E) ----------
    ws.cell(row=1, column=4, value='PON Layers Missing/Present').font = Font(bold=True)
    ws.cell(row=1, column=5, value='Status').font = Font(bold=True)

    patterns = [
        ('Conduit', f"{modules.config.DATA_DIR}/*conduit*.geojson"),
        ('Distribution Aerial', f"{modules.config.DATA_DIR}/*fiber-distribution-aerial*.geojson"),
        ('Distribution Underground', f"{modules.config.DATA_DIR}/*fiber-distribution-underground*.geojson"),
        ('Fiber-Drops', f"{modules.config.DATA_DIR}/*fiber-drop*.geojson"),
        ('NIDs', f"{modules.config.DATA_DIR}/*ni-ds*.geojson"),
        ('NAPs', f"{modules.config.DATA_DIR}/*nap*.geojson"),
        ('Power Poles', f"{modules.config.DATA_DIR}/*power-pole*.geojson"),
        ('Service Locations', f"{modules.config.DATA_DIR}/*service-location*.geojson"),
        ('Slack-Loops', f"{modules.config.DATA_DIR}/*slack-loop*.geojson"),
        ('Vaults', f"{modules.config.DATA_DIR}/*vault*.geojson"),
    ]

    for ridx, (desc, patt) in enumerate(sorted(patterns, key=lambda x: natural_key(x[0])), start=3):
        desc_cell = ws.cell(row=ridx, column=4, value=desc)
        exists = bool(glob.glob(patt))
        symbol = '☑' if exists else '☐'
        status_cell = ws.cell(row=ridx, column=5, value=symbol)
        status_cell.alignment = Alignment(horizontal='center')

        # Green for present, red for missing
        font_color = '008000' if exists else 'FF0000'
        bg_color = 'C6EFCE' if exists else 'FFC7CE'
        status_cell.font = Font(color=font_color)
        status_cell.fill = PatternFill(fill_type='solid', start_color=bg_color)

        # Emphasize missing line
        if not exists:
            desc_cell.font = Font(bold=True)

    # ---------- Finalize ----------
    auto_size(wb)
    apply_borders(ws)


def write_distribution_and_nap_walker_sheet(wb, issues: list[dict]):
    """
    Create an Excel sheet named 'Distribution and NAP Walker' from the issues
    returned by modules.hard_scripts.distribution_walker.find_deep_distribution_mismatches().

    Expected issue keys (any may be absent depending on issue type):
      - path (str)
      - nap_id (str)
      - dist_id (str)
      - svc_id (str)
      - found_drop_color (str)   # for 'Drop color not expected at NAP'
      - drop_color (str)         # for 'Service Location splice color mismatch'
      - svc_colors (list[str])
      - expected_colors (list[str])
      - found_drops (list[dict]) # {drop_id, color, distance_m}
      - missing_colors (list[str])
      - issue (str)
    """
    from openpyxl.styles import Font, Alignment
    import logging
    import modules.config
    from modules.basic.log_configs import format_table_lines

    logger = logging.getLogger(__name__)

    # If there are no issues and we’re not in “show all” mode, do not create the sheet.
    if not issues and not getattr(modules.config, "SHOW_ALL_SHEETS", False):
        return  # ← key change: nothing written, sheet won’t exist

    # 1) Create sheet
    ws = wb.create_sheet(title='Distribution and NAP Walker')
    ws.freeze_panes = 'A2'

    # 2) Header row
    headers = [
        'Path',
        'NAP ID',
        'Dist. ID',
        'Service Location ID',
        'Drop Color',
        'SL Colors',
        'Expected Colors',
        'Missing Colors',
        'Found Drops',
        'Issue',
    ]
    for c, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # 3) Normalize helpers
    def _csv(v):
        if v is None:
            return ''
        if isinstance(v, (list, tuple, set)):
            return ', '.join(str(x) for x in v)
        return str(v)

    def _fmt_found_drops(v):
        if not isinstance(v, (list, tuple)):
            return ''
        out = []
        for d in v:
            if not isinstance(d, dict):
                continue
            did = d.get('drop_id', '')
            col = d.get('color', '')
            dist = d.get('distance_m', '')
            dist_part = f"d={dist}m" if dist != '' else ""
            out.append(f"{did}={col}({dist_part})" if dist_part else f"{did}={col}")
        return ', '.join(out)

    # 4) Rows
    rows_for_log = []
    row_idx = 2
    for it in (issues or []):
        path = it.get('path', '')
        nap_id = it.get('nap_id', '')
        dist_id = it.get('dist_id', '')
        svc_id = it.get('svc_id', '')
        drop_col = it.get('found_drop_color') or it.get('drop_color', '')
        svc_cols = _csv(it.get('svc_colors', []))
        expected = _csv(it.get('expected_colors', []))
        missing = _csv(it.get('missing_colors', []))
        found_dps = _fmt_found_drops(it.get('found_drops', []))
        issue_txt = (it.get('issue') or '').strip()

        row_vals = [path, nap_id, dist_id, svc_id, drop_col, svc_cols, expected, missing, found_dps, issue_txt]
        for c, val in enumerate(row_vals, start=1):
            ws.cell(row=row_idx, column=c, value=val)
        rows_for_log.append([str(x) if x is not None else '' for x in row_vals])
        row_idx += 1

    # 5) Error-only logging (no Excel mirror)
    from modules.basic.log_configs import format_table_lines
    if rows_for_log:
        logger.error(f"==== [Distribution and NAP Walker] Errors ({len(rows_for_log)}) ====")
        for line in format_table_lines(headers, rows_for_log):
            logger.error(f"[Distribution and NAP Walker] {line}")
        logger.info("==== End [Distribution and NAP Walker] Errors ====")


    apply_borders(ws)


def write_geojson_summary(
    ws,
    slack_counter,
    stacked_coords,
    slack_missing,
    nap_map,
    vault_map,
    t3_map,
    power_map
):
    """
    Writes the 'GeoJSON Summary' sheet, including T-3 vault IDs.
    """
    r = 1

    # Metrics rows
    for label, val in [
        ('Total Slack Loop features',     sum(slack_counter.values())),
        ('Unique Slack Loop coords',      len(slack_counter)),
        ('Slack Loop coords stacked',     len(stacked_coords)),
        ('Slack Loops missing nap/vault', len(slack_missing)),
    ]:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=val)
        r += 1

    # blank row
    r += 1

    header_font = Font(bold=True)
    # Stacked Slack Loops header
    c1 = ws.cell(row=r, column=1, value='Stacked Slack Loops:')
    c1.font = header_font
    r += 1
    
    headers = ['Coordinate', 'Count', 'NAP, Vault, or Pole ID']
    header_font = Font(bold=True)

    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=col_idx, value=title)
        cell.font = header_font

    r += 1
# Stacked Slack Loops data (now sorted by count desc, then NAP ID naturally)
    for coord, cnt in sorted(
        stacked_coords.items(),
        key=lambda item: (
            -item[1],  # 1) highest counts first
            natural_key(
                # 2) tie-break by whichever ID you pulled for this coord
                t3_map.get(item[0])
                or nap_map.get(item[0])
                or vault_map.get(item[0])
                or ''
            )
        )
    ):
        # write coords
        ws.cell(row=r, column=1, value=f'{coord[0]:.6f}, {coord[1]:.6f}')
        # write count
        ws.cell(row=r, column=2, value=cnt)
        # pick ID in priority order
        # collect any matching IDs: T-3 vault, NAP, Vault, then Power‐Pole
        ids = []

        if coord in t3_map:
            ids.append(t3_map[coord])
        if coord in nap_map:
            ids.append(nap_map[coord])
        # only include Vault IDs when this coord isn’t already a T-3 Vault, NAP or Power Pole
        if (
            coord in vault_map
            and coord not in t3_map
            and coord not in nap_map
            and coord not in power_map
        ):
            ids.append(vault_map[coord])
        if coord in power_map:
            ids.append(power_map[coord])
        # join them with commas, so pole IDs appear alongside NAP/Vault IDs
        ws.cell(row=r, column=3, value=', '.join(ids))
        r += 1

    # Only show "Missing Slack Loops" if there are missing points
    # or if the user has enabled SHOW_ALL_SHEETS in config
    if slack_missing or modules.config.SHOW_ALL_SHEETS:
        r += 1
        ws.cell(row=r, column=1, value='Missing Slack Loops:')
        r += 1

        for coord in sorted(slack_missing):
            ws.cell(row=r, column=1, value=f'{coord[0]:.6f}, {coord[1]:.6f}')
            r += 1
    apply_borders(ws)



def write_person_sheets(wb, df: pd.DataFrame, patterns: list, id_col: str):
    """
    Creates one sheet per layer in df['Layer Name'], with
    person-by-pattern Yes/No and full-text examples + NAP stats.
    """
    persons = df['Edited By'].replace('', 'Unknown').unique()
    for layer in df['Layer Name'].unique():
        sub = df[df['Layer Name'] == layer]
        pats = [p for p in patterns if sub['Type of Edit'].str.contains(p, na=False).any()]
        if not pats:
            continue

        ws = wb.create_sheet(title=layer[:31] or layer)
        # Header
        ws.cell(row=1, column=1, value='Person')
        for c, pat in enumerate(pats, start=2):
            ws.cell(row=1, column=c, value=pat)
        # Data rows
        for r_idx, person in enumerate(persons, start=2):
            ws.cell(row=r_idx, column=1, value=person)
            usr = sub[sub['Edited By'] == person]
            for c, pat in enumerate(pats, start=2):
                ws.cell(
                    row=r_idx,
                    column=c,
                    value='Yes' if usr['Type of Edit'].str.contains(pat, na=False).any() else 'No'
                )
        # Full-text examples + NAP stats
        bot = len(persons) + 3
        for pat in pats:
            ws.cell(row=bot, column=1, value=f'{pat}:')
            bot += 1
            for full in sub.loc[sub['Type of Edit'].str.contains(pat, na=False), 'Type of Edit'].unique():
                ws.cell(row=bot, column=1, value=full)
                bot += 1
            bot += 1

        if layer.lower() == 'nap':
            m = sub[id_col].str.match(r'^SC-\d+$', na=False)
            created = sub.loc[(sub['Type of Edit'] == 'Feature was created') & m, id_col].unique().tolist()
            deleted = sub.loc[(sub['Type of Edit'] == 'Feature was deleted') & m, id_col].unique().tolist()
            bot += 1
            ws.cell(row=bot, column=1, value='--- NAP Feature stats ---'); bot += 1
            ws.cell(row=bot, column=1, value='Created count'); ws.cell(row=bot, column=2, value=len(created)); bot += 1
            ws.cell(row=bot, column=1, value='Deleted count'); ws.cell(row=bot, column=2, value=len(deleted)); bot += 2
            ws.cell(row=bot, column=1, value='Created IDs:'); bot += 1
            for cid in created:
                ws.cell(row=bot, column=2, value=cid); bot += 1
            bot += 1
            ws.cell(row=bot, column=1, value='Deleted IDs:'); bot += 1
            for did in deleted:
                ws.cell(row=bot, column=2, value=did); bot += 1
            sc, sd = set(created), set(deleted)
            if sc == sd:
                ws.cell(row=bot, column=1, value='✅ All created/deleted IDs match')
            else:
                if diff := sorted(sd - sc):
                    ws.cell(row=bot, column=1, value='IDs deleted but not created:')
                    ws.cell(row=bot, column=2, value=', '.join(diff)); bot += 1
                if diff := sorted(sc - sd):
                    ws.cell(row=bot, column=1, value='IDs created but not deleted:')
                    ws.cell(row=bot, column=2, value=', '.join(diff))
    apply_borders(ws)


# DON'T THINK I NEED THIS ANYMORE... BUT I'LL WAIT.
# def write_missing_slack_sheet(wb, missing_poles: list, power_map: dict):
#     """
#     Writes the 'Fiber Missing Slack Loops' sheet.
#     """
#     ws = wb.create_sheet(title='Fiber Missing Slack Loops')
#     ws.cell(row=1, column=1, value='Power Pole ID')
#     ws.cell(row=1, column=2, value='Latitude')
#     ws.cell(row=1, column=3, value='Longitude')
#     for r, (lat, lon) in enumerate(missing_poles, start=2):
#         ws.cell(row=r, column=1, value=power_map.get((round(lat,6), round(lon,6)), ''))
#         ws.cell(row=r, column=2, value=f'{lat:.6f}')
#         ws.cell(row=r, column=3, value=f'{lon:.6f}')


def write_drop_issues_sheet(wb, mismatches_or_service_coords, drop_coords=None, combined=None):
    """
    'Drop Issues' sheet.

    Backwards-compatible call styles:
      • New (current repo): write_drop_issues_sheet(wb, mismatches)
      • Legacy (your local main.py): write_drop_issues_sheet(wb, service_coords, drop_coords, combined)

    Only the IDs in `mismatches` / `combined` are used to render this sheet.
    """
    from openpyxl.styles import Font, Alignment
    import logging
    import modules.config
    from modules.basic.log_configs import format_table_lines

    logger = logging.getLogger(__name__)

    # --- Back-compat adapter ---
    # If called with 4 args, `combined` is the list/dict we actually need.
    mismatches = combined if combined is not None else mismatches_or_service_coords

    # 1) Create sheet
    ws = wb.create_sheet(title='Drop Issues')
    ws.freeze_panes = 'A5'

    # 2) Banner (A1:C3)
    ws.merge_cells('A1:C3')
    header_text = (
        'Missing Attributes on Service Locations and/or wrong Drop Color going to '
        'Service Locations - If Errors still happen, check if the color is the '
        'correct color, not others like "Purple".'
    )
    header_cell = ws['A1']
    header_cell.value = header_text
    header_cell.alignment = Alignment(horizontal='center', wrap_text=True)
    header_cell.font = Font(bold=True)

    # 3) Table column titles on row 4 (add Issue column)
    headers = ['Service Location ID', 'Missing Drops (Service Location ID)', 'Issue']
    for c, title in enumerate(headers, start=1):
        ws.cell(row=4, column=c, value=title)

    # 4) Normalize 'mismatches' to a mapping or ordered list of SIDs
    # Accepts: list/tuple/set of SIDs or dict {sid: missing_sid}
    if isinstance(mismatches, dict):
        ordered_sids = sorted(mismatches.keys())
        getter = mismatches.get
    else:
        ordered_sids = sorted(mismatches or [])
        getter = lambda _sid: None  # noqa: E731

    # 5) Write rows (starting at row 5)
    rows_written = []
    for idx, sid in enumerate(ordered_sids, start=5):
        miss_val = getter(sid)
        ws.cell(row=idx, column=1, value=sid)
        if miss_val:
            ws.cell(row=idx, column=2, value=miss_val)
        # New Issue column with short description
        issue_text = "Missing drop or color mismatch"
        ws.cell(row=idx, column=3, value=issue_text)
        rows_written.append([sid, miss_val, issue_text])

    # 6) Error-only logging (no Excel mirror)
    if rows_written:
        logger.error(f"==== [Drop Issues] Errors ({len(rows_written)}) ====")
        for line in format_table_lines(headers, rows_written):
            logger.error(f"[Drop Issues] {line}")
        logger.info("==== End [Drop Issues] Errors ====")

    # 7) Borders
    apply_borders(ws)


# def write_drop_issues_sheet(wb, mismatches):
#     """
#     'Drop Issues' sheet. Shows Service Locations that are missing drops or have a
#     color mismatch. Adds an 'Issue' column. Also logs the same errors in the log
#     (no Excel→log mirror; error-only logging instead).
#     """
#     from openpyxl.styles import Font, Alignment
#     import logging
#     import modules.config
#     from modules.basic.log_configs import format_table_lines

#     logger = logging.getLogger(__name__)

#     # 1) Create sheet
#     ws = wb.create_sheet(title='Drop Issues')
#     ws.freeze_panes = 'A5'

#     # 2) Banner (A1:C3)
#     ws.merge_cells('A1:C3')
#     header_text = (
#         'Missing Attributes on Service Locations and/or wrong Drop Color going to '
#         'Service Locations - If Errors still happen, check if the color is the '
#         'correct color, not others like "Purple".'
#     )
#     header_cell = ws['A1']
#     header_cell.value = header_text
#     header_cell.alignment = Alignment(horizontal='center', wrap_text=True)
#     header_cell.font = Font(bold=True)

#     # 3) Table column titles on row 4 (add Issue column)
#     headers = ['Service Location ID', 'Missing Drops (Service Location ID)', 'Issue']
#     for c, title in enumerate(headers, start=1):
#         ws.cell(row=4, column=c, value=title)

#     # 4) Normalize 'mismatches' to a mapping or ordered list of SIDs
#     # Accepts: list/tuple/set of SIDs or dict {sid: missing_sid}
#     if isinstance(mismatches, dict):
#         ordered_sids = sorted(mismatches.keys())
#         getter = mismatches.get
#     else:
#         ordered_sids = sorted(mismatches or [])
#         getter = lambda _sid: None  # noqa: E731

#     # 5) Write rows (starting at row 5)
#     rows_written = []
#     for idx, sid in enumerate(ordered_sids, start=5):
#         miss_val = getter(sid)
#         ws.cell(row=idx, column=1, value=sid)
#         if miss_val:
#             ws.cell(row=idx, column=2, value=miss_val)
#         # New Issue column with short description
#         issue_text = "Missing drop or color mismatch"
#         ws.cell(row=idx, column=3, value=issue_text)
#         rows_written.append([sid, miss_val, issue_text])

#     # 6) Error-only logging (no Excel mirror)
#     if rows_written:
#         logger.error(f"==== [Drop Issues] Errors ({len(rows_written)}) ====")
#         for line in format_table_lines(headers, rows_written):
#             logger.error(f"[Drop Issues] {line}")
#         logger.info("==== End [Drop Issues] Errors ====")

#     # 7) Borders
#     apply_borders(ws)



def write_slack_loop_issues_sheet(wb, sd_issues, ug_issues, aerial_issues=None, tail_issues=None):
    """
    Slack Loop Issues sheet with a horizontal summary bar and side-by-side detail blocks.

    Sections (each shown only if it has rows OR modules.config.SHOW_ALL_SHEETS):
      • Slack-Distribution mismatches (Distribution)  — tuples: (slack_vid, fiber_label, dist_ID, issue)
      • Underground Slack Loop issues (Underground)   — tuples: (touching_dist_ids, "underground",
                                                                existing_slack_labels, slack_loop_labels,
                                                                slack_loop_vids, vault_or_nap_vid, issue)
      • Aerial Slack Loop issues (Aerial)             — tuples: (power_pole_id, latitude, longitude, issue)
      • Distribution End Tail issues (Tail End)       — tuples: (slack_loop_vid, type, slack_label, expected_label)
    """
    from openpyxl.styles import Font, Alignment
    import modules.config

    ws = wb.create_sheet(title='Slack Loop Issues', index=2)
    show_all = bool(getattr(modules.config, "SHOW_ALL_SHEETS", False))

    # ───────────────────────────── Summary bar (rows 1–3) ─────────────────────────────
    labels = ["Distribution", "Underground", "Aerial", "Tail End"]
    counts = [
        len(sd_issues or []),
        len(ug_issues or []),
        len(aerial_issues or []),
        len(tail_issues or []),
    ]

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(labels))
    t = ws.cell(row=1, column=1, value="Slack Loop Issues — Summary")
    t.font = Font(bold=True)
    t.alignment = Alignment(horizontal="center")

    # Labels (row 2) and values (row 3)
    for c, lab in enumerate(labels, start=1):
        cell = ws.cell(row=2, column=c, value=lab)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for c, val in enumerate(counts, start=1):
        cell = ws.cell(row=3, column=c, value=val)
        cell.alignment = Alignment(horizontal="center")
        if isinstance(val, int) and val > 0:
            cell.font = Font(bold=True)

    # Leave row 4 blank (visual spacer to match your screenshot)

    # Start detail blocks at row 5; freeze rows 1–6 so headers stay put.
    DETAIL_START_ROW = 5
    FREEZE_ROW = 7
    ws.freeze_panes = f"A{FREEZE_ROW}"

    # Helper to join list/tuple values for cells
    def _join(v):
        if isinstance(v, (list, tuple, set)):
            return ", ".join(str(x) for x in v)
        return "" if v is None else str(v)

    # ─────────────────────── Layout plan: side-by-side blocks ───────────────────────
    blocks = []

    # Block A: Distribution (Slack-Distribution mismatches)
    if (sd_issues and len(sd_issues) > 0) or show_all:
        blocks.append({
            "title": "Slack-Distribution Mismatches",
            "headers": ["Slack Vetro ID", "Fiber Label", "Distribution ID", "Issue"],
            "rows": [[_join(a), _join(b), _join(c), _join(d)] for (a, b, c, d) in (sd_issues or [])],
            "width": 4,
        })

    # Block B: Underground
    if (ug_issues and len(ug_issues) > 0) or show_all:
        blocks.append({
            "title": "Underground Slack Loop Issues (Vault/NAP anchors)",
            "headers": [
                "Touching Distribution ID(s)",
                "Type",
                "Existing Slack Fiber Label(s)",
                "Slack Loop Label(s)",
                "Slack Loop Vetro ID(s)",
                "Vault/NAP Vetro ID",
                "Issue",
            ],
            "rows": [
                [
                    _join(t0), _join(t1), _join(t2),
                    _join(t3), _join(t4), _join(t5), _join(t6),
                ]
                for (t0, t1, t2, t3, t4, t5, t6) in (ug_issues or [])
            ],
            "width": 7,
        })

    # Block C: Aerial
    if (aerial_issues and len(aerial_issues) > 0) or show_all:
        blocks.append({
            "title": "Aerial Slack Loop Issues (Pole has Drop + NAP but no Slack)",
            "headers": ["Power Pole ID", "Latitude", "Longitude", "Issue"],
            "rows": [[_join(pid), _join(lat), _join(lon), _join(issue)] for (pid, lat, lon, issue) in (aerial_issues or [])],
            "width": 4,
        })


    # Block D: Tail End (Distribution End Tail issues)
    if (tail_issues and len(tail_issues) > 0) or show_all:
        # Detect if we have any "No slack..." rows; only then add the extra column.
        has_no_slack = any(
            (str(issue or "").lower().startswith("no slack loop near distribution end"))
            for (_vid, _kind, _lbl, _exp, issue) in (tail_issues or [])
        )

        headers = [
            "Slack Loop ID", "Type", "Slack Loop", "Expected Slack Loop", "Issue",
        ]

        # Try to pull the sidecar Distribution IDs aligned to tail_issues from slack_loops
        sidecar = None
        if has_no_slack:
            try:
                from modules.simple_scripts import slack_loops as _slack_loops_mod
                sidecar = getattr(_slack_loops_mod, "_LAST_TAIL_END_DIST_IDS", None)
                # Basic sanity: must be same length to rely on alignment
                if sidecar is not None and len(sidecar) != len(tail_issues):
                    sidecar = None
            except Exception:
                sidecar = None

        rows = []
        if has_no_slack:
            headers.append("Distribution ID")
            for idx, (vid, kind, lbl, exp, issue) in enumerate(tail_issues or []):
                dist_id = ""
                if sidecar is not None:
                    dist_id = sidecar[idx] if str(issue or "").lower().startswith("no slack loop near distribution end") else ""
                rows.append([_join(vid), _join(kind), _join(lbl), _join(exp), _join(issue), _join(dist_id)])
            width = 6
        else:
            rows = [
                [_join(vid), _join(kind), _join(lbl), _join(exp), _join(issue)]
                for (vid, kind, lbl, exp, issue) in (tail_issues or [])
            ]
            width = 5

        blocks.append({
            "title": "Distribution End Tail Issues",
            "headers": headers,
            "rows": rows,
            "width": width,
        })

    if not blocks:
        blocks = [{
            "title": "Slack-Distribution Mismatches",
            "headers": ["Slack Vetro ID", "Fiber Label", "Distribution ID", "Issue"],
            "rows": [],
            "width": 4,
        }]

    # ───────────────────────── Render blocks side-by-side ─────────────────────────
    current_col = 1
    for blk in blocks:
        title = blk["title"]
        headers = blk["headers"]
        rows = blk["rows"]
        width = blk["width"]

        # Title cell for this block
        tcell = ws.cell(row=DETAIL_START_ROW, column=current_col, value=title)
        tcell.font = Font(bold=True)
        tcell.alignment = Alignment(horizontal="center")
        ws.merge_cells(
            start_row=DETAIL_START_ROW,
            start_column=current_col,
            end_row=DETAIL_START_ROW,
            end_column=current_col + width - 1
        )

        # Header row (row 6)
        for i, h in enumerate(headers, start=current_col):
            hc = ws.cell(row=DETAIL_START_ROW + 1, column=i, value=h)
            hc.font = Font(bold=True)
            hc.alignment = Alignment(horizontal="center")

        # Data rows (begin at row 7)
        r = DETAIL_START_ROW + 2
        for row_vals in rows:
            for i, val in enumerate(row_vals, start=current_col):
                ws.cell(row=r, column=i, value=val)
            r += 1

        # Advance to next block (1-col gutter)
        current_col += width + 1
    apply_borders(ws)


def write_footage_issues_sheet(wb, mismatches):
    """
    'Footage Issues' sheet — renders only the blocks that have rows:
      • Distribution Footage — Missing/Invalid Note
      • Fiber Drops > 250 ft

    If one side is empty (and SHOW_ALL_SHEETS is False), it is not drawn;
    the remaining block starts at column A.
    """
    from openpyxl.styles import Alignment, Font
    from modules.simple_scripts.footage_issues import find_overlength_drop_cables
    import modules.config

    ws = wb.create_sheet(title='Footage Issues')
    ws.freeze_panes = "A3"

    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    show_all = bool(getattr(modules.config, "SHOW_ALL_SHEETS", False))

    # Left block (Distribution note issues)
    left_title = "Distribution Footage Length on Notes field — Missing/Invalid Note"
    left_headers = ["Distribution ID", "Type", "Vetro ID", "Issue"]
    left_rows = []
    for dist_id, kind, vetro_id in (mismatches or []):
        k = (kind or "").lower()
        type_str = "Aerial" if "aerial" in k else ("Underground" if "underground" in k else "")
        left_rows.append([str(dist_id or ""), type_str, str(vetro_id or ""), 'Missing or invalid "Note" footage value'])

    # Right block (Overlength cable > 250 ft)
    right_title = "Fiber Drops > 250 ft"
    right_headers = ["Vetro ID", "Type", "Length (ft)", "Issues"]
    right_rows = []
    for vetro_id, type_str, total_len in (find_overlength_drop_cables(limit_ft=250.0) or []):
        val = f"{float(total_len):.2f}" if isinstance(total_len, (int, float)) else str(total_len or "")
        right_rows.append([str(vetro_id or ""), str(type_str or ""), val, "Over 250 ft"])

    # Choose which blocks to render
    blocks = []
    if left_rows or show_all:
        blocks.append((left_title, left_headers, left_rows))
    if right_rows or show_all:
        blocks.append((right_title, right_headers, right_rows))

    # If nothing to show (and not SHOW_ALL), remove the sheet and exit
    if not blocks and not show_all:
        wb.remove(ws)
        return

    # Render the selected blocks side-by-side
    GAP = 1
    col = 1
    for title, headers, rows in blocks:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + len(headers) - 1)
        th = ws.cell(row=1, column=col, value=title); th.font = bold; th.alignment = center
        for i, h in enumerate(headers, start=col):
            hc = ws.cell(row=2, column=i, value=h); hc.font = bold; hc.alignment = center
        r = 3
        for row_vals in rows:
            for i, val in enumerate(row_vals, start=col):
                ws.cell(row=r, column=i, value=val)
            r += 1
        col += len(headers) + GAP

    apply_borders(ws)


def write_nid_issues(wb, nid_issues: list):
    from openpyxl.styles import Alignment, Font
    import logging
    import modules.config
    from modules.basic.log_configs import format_table_lines

    logger = logging.getLogger(__name__)

    # Create sheet
    ws = wb.create_sheet(title='NID Issues')

    # Banner across 7 columns
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    banner = ws.cell(
        row=1, column=1,
        value='Checks if Splice Colors mismatch; Service Locations have correct 1.1, 1.2, etc. attributes.'
    )
    banner.alignment = Alignment(horizontal='center')
    banner.font = Font(bold=True)

    # Column headers (row 2)
    headers = [
        'NID ID',
        'Issue',
        'Service Location ID',
        'Service Location Color',
        'Drop Color',
        'Expected Splice',
        'Actual Splice',
    ]
    for c, t in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=c, value=t)
        cell.font = Font(bold=True)

    # Freeze just the first row (banner)
    ws.freeze_panes = 'A3'

    # Data (from row 3)
    for r, issue in enumerate(nid_issues or [], start=3):
        ws.cell(row=r, column=1, value=issue.get('nid'))
        ws.cell(row=r, column=2, value=issue.get('issue', ''))
        ws.cell(row=r, column=3, value=issue.get('svc_id', ''))
        ws.cell(row=r, column=4, value=issue.get('svc_color', ''))
        ws.cell(row=r, column=5, value=issue.get('drop_color', ''))
        ws.cell(row=r, column=6, value=issue.get('expected_splice', ''))
        ws.cell(row=r, column=7, value=issue.get('actual_splice', ''))

    # Center columns 4..7 (headers + data)
    center = Alignment(horizontal='center')
    for col in range(4, ws.max_column + 1):
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col).alignment = center

    # ---------------------------
    # Error-only logging (single)
    # ---------------------------

    # Emoji color helpers (used only for log prettiness when LOG_COLOR_MODE == 'EMOJI')
    def _color_emoji(name: str) -> str:
        mapping = {
            "Blue": "", "Orange": "", "Green": "", "Brown": "",
            "Slate": "◾️", "White": "⬜", "Red": "", "Black": "⬛",
            "Yellow": "", "Violet": "", "Rose": "", "Aqua": "",
        }
        return mapping.get((name or "").strip(), "◻️")

    def _maybe_emojiize_csv(csv_text: str) -> str:
        mode = str(getattr(modules.config, "LOG_COLOR_MODE", "OFF")).upper()
        parts = [p.strip() for p in (csv_text or "").split(",") if p.strip()]
        if mode == "EMOJI":
            return ", ".join(_color_emoji(p) for p in parts)
        return ", ".join(parts)

    def _maybe_emojiize_one(name: str) -> str:
        mode = str(getattr(modules.config, "LOG_COLOR_MODE", "OFF")).upper()
        return _color_emoji(name) if mode == "EMOJI" else (name or "")

    # Prefer the full computed list (OK+errors) if available; otherwise use nid_issues
    rows_source = None
    try:
        from modules.simple_scripts.nids import iterate_nid_checks
        rows_source = iterate_nid_checks(include_ok=True)
    except Exception:
        rows_source = None
    if not rows_source:
        rows_source = nid_issues or []

    # Collect only the errors
    error_rows = []
    for d in rows_source:
        issue_text = (d.get("issue") or "").strip()
        if not issue_text:
            continue
        svc_color_csv = d.get("svc_color") or ""
        drop_color    = d.get("drop_color") or ""
        expected_s    = d.get("expected_splice") or ""
        actual_s      = d.get("actual_splice") or ""

        error_rows.append([
            d.get("nid", ""),
            f"❌ {issue_text}",
            d.get("svc_id") or "(none)",
            _maybe_emojiize_csv(svc_color_csv),
            _maybe_emojiize_one(drop_color),
            _maybe_emojiize_one(expected_s),
            _maybe_emojiize_one(actual_s),
        ])

    if error_rows:
        logger.error(f"==== [NID Issues] Errors ({len(error_rows)}) ====")
        for line in format_table_lines(headers, error_rows):
            logger.error(f"[NID Issues] {line}")
        logger.info("==== End [NID Issues] Errors ====")

    apply_borders(ws)


def write_service_location_attr_issues(wb, records):
    """
    Service Location Issues sheet.
    Layout (unchanged from your reverted version), with:
      • Issue column text = "Missing Attribute"
      • Attribute cells show "Missing" or "✅"
      • NEW: freeze header row; center columns starting at col 3
      • Logging mirror restored
    """
    from openpyxl.styles import Font, Alignment
    import logging, modules.config

    logger = logging.getLogger(__name__)
    ws = wb.create_sheet(title='Service Location Issues')

    # Required attributes (same list you were using)
    attrs = [
        "Build Type", "Building Type", "Drop Type",
        "NAP #", "NAP Location", "Loose Tube", "Splice Colors",
    ]

    # Build SL → per-attr status map
    sl_map: dict[str, dict[str, str]] = {}
    for rec in (records or []):
        sl_id = rec.get("Service Location ID")
        if not sl_id:
            continue
        if sl_id not in sl_map:
            sl_map[sl_id] = {a: "✅" for a in attrs}
        a = rec.get("Attribute")
        if a in sl_map[sl_id]:
            sl_map[sl_id][a] = "Missing"

    # Only include columns that are missing for someone (unless SHOW_ALL_SHEETS)
    if getattr(modules.config, "SHOW_ALL_SHEETS", False):
        missing_cols = attrs
    else:
        missing_cols = [a for a in attrs if any(flags[a] == "Missing" for flags in sl_map.values())]

    # Header
    headers = ["Service Location ID", "Issue"] + missing_cols
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Rows
    rows_written = []
    for sl_id, flags in sl_map.items():
        row = [sl_id, "Missing Attribute"] + [("Missing" if flags.get(a) == "Missing" else "✅") for a in missing_cols]
        ws.append(row)
        rows_written.append(row)

    # Center columns starting at column 3 (C…)
    center = Alignment(horizontal="center")
    for col in range(3, ws.max_column + 1):
        for r in range(1, ws.max_row + 1):
            ws.cell(row=r, column=col).alignment = center

    # Logging mirror (restored)
    if getattr(modules.config, "LOG_MIRROR_SHEETS", False) and getattr(modules.config, "LOG_SVCLOC_SHEET_TO_LOG", False):       
        logger.info(" | ".join(headers))
        for row in rows_written:
            logger.error(" | ".join(str(v) if v is not None else "" for v in row))
    apply_borders(ws)


def write_nap_issues_sheet(wb, nap_mismatches, id_format_issues):
    """
    NAP Issues sheet — renders only the blocks that have rows:
      • NAP Mismatches
      • NAP Naming Issues
      • Warnings (NAP Specs)

    If a block has no rows (and SHOW_ALL_SHEETS is False), it is omitted entirely.
    Blocks are laid out left→right with a 1-column gutter between them.
    """
    from openpyxl.styles import Font, Alignment
    from modules.simple_scripts.nap_rules import scan_nap_spec_warnings
    import logging, modules.config

    logger = logging.getLogger(__name__)
    ws = wb.create_sheet(title="NAP Issues")
    ws.freeze_panes = "A3"  # lock headers
    bold = Font(bold=True)
    center = Alignment(horizontal='center')

    def _join(v):
        if v is None:
            return ""
        if isinstance(v, (list, tuple, set)):
            return ", ".join(str(x) for x in v)
        return str(v)

    show_all = bool(getattr(modules.config, "SHOW_ALL_SHEETS", False))

    # --- Build rows for each logical block ---
    # A) NAP Mismatches
    a_headers = ["NAP ID", "Loose Tube", "Missing Indices", "Missing Colors", "Issue"]
    a_rows = []
    for rec in (nap_mismatches or []):
        if isinstance(rec, dict):
            nap = rec.get("nap") or rec.get("NAP ID") or ""
            loose = rec.get("loose_abbrev") or rec.get("Loose Tube") or ""
            miss_idx = rec.get("missing_indices") or rec.get("Missing Indices") or []
            miss_col = rec.get("missing_colors") or rec.get("Missing Colors") or []
        else:
            nap = rec[0] if isinstance(rec, (list, tuple)) and len(rec) > 0 else ""
            loose = rec[1] if isinstance(rec, (list, tuple)) and len(rec) > 1 else ""
            miss_idx = rec[2] if isinstance(rec, (list, tuple)) and len(rec) > 2 else []
            miss_col = rec[3] if isinstance(rec, (list, tuple)) and len(rec) > 3 else []
        a_rows.append([nap, loose, _join(miss_idx), _join(miss_col), "Loose-tube color mismatch"])

    # B) NAP Naming Issues
    b_headers = ["NAP", "Vetro ID", "Issue"]
    b_rows = []
    for rec in (id_format_issues or []):
        if isinstance(rec, (list, tuple)):
            nap_id = rec[0] if len(rec) > 0 else ""
            vetro_id = rec[1] if len(rec) > 1 else ""
        else:
            nap_id = rec.get("nap_id") or rec.get("nap") or rec.get("NAP") or ""
            vetro_id = rec.get("vetro_id") or rec.get("Vetro ID") or ""
        b_rows.append([nap_id, vetro_id, "NAP ID format issue"])

    # C) Spec warnings (computed within this writer)
    c_headers = ["NAP ID", "Field", "Value", "Hint", "Issue"]
    c_rows = []
    for rec in scan_nap_spec_warnings() or []:
        c_rows.append([
            rec.get("NAP ID", "") or rec.get("nap_id", ""),
            rec.get("Field", "") or rec.get("field", ""),
            rec.get("Value", "") or rec.get("value", ""),
            rec.get("Hint", "") or rec.get("hint", ""),
            "Spec warning",
        ])

    # --- Collect blocks that should actually render ---
    blocks = []
    if a_rows or show_all:
        blocks.append(("NAP Mismatches", a_headers, a_rows))
    if b_rows or show_all:
        blocks.append(("NAP Naming Issues", b_headers, b_rows))
    if c_rows or show_all:
        blocks.append(("Warnings (NAP Specs)", c_headers, c_rows))

    # If there’s truly nothing to show (and not SHOW_ALL), drop the sheet.
    if not blocks and not show_all:
        wb.remove(ws)
        return

    # --- Render the chosen blocks side-by-side ---
    GAP = 1
    col = 1
    for title, headers, rows in blocks:
        # Title row (1), merged to the width of headers
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + len(headers) - 1)
        tcell = ws.cell(row=1, column=col, value=title)
        tcell.font = bold
        tcell.alignment = center

        # Header row (2)
        for i, h in enumerate(headers, start=col):
            hc = ws.cell(row=2, column=i, value=h)
            hc.font = bold
            hc.alignment = center

        # Data starting at row 3
        r = 3
        for row_vals in rows:
            for i, val in enumerate(row_vals, start=col):
                ws.cell(row=r, column=i, value=val)
            r += 1

        col += len(headers) + GAP

    from modules.basic.log_configs import format_table_lines

    # Error-only logging for each NAP block (no Excel mirror)
    if a_rows:
        logger.error(f"==== [NAP Mismatches] Errors ({len(a_rows)}) ====")
        for line in format_table_lines(a_headers, a_rows):
            logger.error(f"[NAP Mismatches] {line}")
        logger.info("==== End [NAP Mismatches] Errors ====")

    if b_rows:
        logger.error(f"==== [NAP Naming Issues] Errors ({len(b_rows)}) ====")
        for line in format_table_lines(b_headers, b_rows):
            logger.error(f"[NAP Naming Issues] {line}")
        logger.info("==== End [NAP Naming Issues] Errors ====")

    if c_rows:
        logger.error(f"==== [Warnings (NAP Specs)] Errors ({len(c_rows)}) ====")
        for line in format_table_lines(c_headers, c_rows):
            logger.error(f"[Warnings (NAP Specs)] {line}")
        logger.info("==== End [Warnings (NAP Specs)] Errors ====")


    apply_borders(ws)


def write_power_pole_issues_sheet(wb, issues: list[dict]):
    """
    Writes the 'Power Pole Issues' sheet for any bend ≥ threshold without anchor,
    with a WIP banner header.
    """
    from openpyxl.styles import Font, Alignment

    ws = wb.create_sheet(title="Power Pole Issues")
    ws.freeze_panes = 'A5'

    # 1) manual widths to constrain A–C
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 48
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 28  # Issue column

    # 2) banner (rows 1–3)
    ws.merge_cells('A1:E3')
    banner = ws['A1']
    banner.value = (
        "⚠️ WIP — Identifies aerial distribution bends ≥ threshold that appear to lack a pole anchor."
    )
    banner.font = Font(bold=True)
    banner.alignment = Alignment(horizontal='center', wrap_text=True)

    # 3) headers (row 4) — add Issue column
    titles = ["Power Pole ID", "Distribution ID", "Bend Angle (°)", "Note", "Issue"]
    for idx, txt in enumerate(titles, start=1):
        c = ws.cell(row=4, column=idx, value=txt)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center')

    # 4) data starting at row 5
    for r, issue in enumerate(issues, start=5):
        ws.cell(row=r, column=1, value=issue["pole_id"])
        ws.cell(row=r, column=2, value=issue["dist_id"])
        ws.cell(row=r, column=3, value=issue["angle"])
        ws.cell(row=r, column=4, value=issue.get("note",""))
        ws.cell(row=r, column=5, value="Unanchored bend ≥ threshold")
    apply_borders(ws)


def write_conduit_sheet(wb, results: dict):
    """
    'Conduit' sheet — renders only the blocks that have rows:
      • Distribution Without Conduit
      • Conduits Without Distribution
      • Conduit Type Issues
    """
    from openpyxl.styles import Alignment, Font
    import modules.config

    ws = wb.create_sheet(title='Conduit')
    ws.freeze_panes = 'A3'

    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    show_all = bool(getattr(modules.config, "SHOW_ALL_SHEETS", False))

    # Lists (default to [])
    df_missing = results.get('df_missing_conduit', []) or []
    cd_missing = results.get('conduit_missing_distribution', []) or []
    type_issues = results.get('type_issues', []) or []

    # Define blocks
    left_title = "Distribution Without Conduit"
    left_headers = ["Distribution ID", "Vetro ID", "Issue"]
    left_rows = [[str(r.get("Distribution ID","")), str(r.get("Vetro ID","")), str(r.get("Issue",""))] for r in df_missing]

    mid_title = "Conduits Without Distribution"
    mid_headers = ["Conduit ID", "Conduit Vetro ID", "Issue"]
    mid_rows = [[str(r.get("Conduit ID","")), str(r.get("Conduit Vetro ID","")), str(r.get("Issue",""))] for r in cd_missing]

    right_title = "Conduit Type Issues"
    right_headers = ["Conduit ID", "Conduit Vetro ID", "Conduit Type", "Issue"]
    right_rows = [[str(r.get("Conduit ID","")), str(r.get("Conduit Vetro ID","")), str(r.get("Conduit Type","")), str(r.get("Issue",""))] for r in type_issues]

    blocks = []
    if left_rows or show_all:
        blocks.append((left_title, left_headers, left_rows))
    if mid_rows or show_all:
        blocks.append((mid_title, mid_headers, mid_rows))
    if right_rows or show_all:
        blocks.append((right_title, right_headers, right_rows))

    # If nothing to show (and not SHOW_ALL), remove the sheet
    if not blocks and not show_all:
        wb.remove(ws)
        return

    # Render selected blocks side-by-side
    GAP = 1
    col = 1
    for title, headers, rows in blocks:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + len(headers) - 1)
        th = ws.cell(row=1, column=col, value=title); th.font = bold; th.alignment = center
        for i, h in enumerate(headers, start=col):
            hc = ws.cell(row=2, column=i, value=h); hc.font = bold; hc.alignment = center
        r = 3
        for row_vals in rows:
            for i, val in enumerate(row_vals, start=col):
                ws.cell(row=r, column=i, value=val)
            r += 1
        col += len(headers) + GAP

    apply_borders(ws)


def write_vaults_sheet(wb, results: dict):
    """
    'Vaults' sheet — renders only the blocks that have rows:
      • Vaults Missing Conduit
      • Vault Spacing > 500 ft
      • Sharp Bends Without Nearby Vault (<130°, >300 ft)
    """
    from openpyxl.styles import Alignment, Font
    import modules.config

    def _to_num(x):
        try:
            return float(x)
        except Exception:
            s = str(x).strip()
            try:
                return float(s.replace(",", ""))
            except Exception:
                return x

    ws = wb.create_sheet(title='Vaults')
    ws.freeze_panes = 'A3'

    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    show_all = bool(getattr(modules.config, "SHOW_ALL_SHEETS", False))

    missing = results.get('vaults_missing_conduit', []) or []
    spacing = results.get('vault_spacing_issues', []) or []
    bends = results.get('bend_vault_issues', []) or []

    left_title = "Vaults Missing Conduit"
    left_headers = ["Vault Vetro ID", "Issue"]
    left_rows = [[str(r.get("Vault Vetro ID","")), str(r.get("Issue",""))] for r in missing]

    mid_title = "Vault Spacing > 500 ft"
    mid_headers = ["Conduit ID", "Conduit Vetro ID", "From Vault", "To Vault", "Distance (ft)", "Limit (ft)", "Issue"]
    mid_rows = [[
        str(r.get("Conduit ID","")),
        str(r.get("Conduit Vetro ID","")),
        str(r.get("From Vault","")),
        str(r.get("To Vault","")),
        _to_num(r.get("Distance (ft)","")),
        _to_num(r.get("Limit (ft)","")),
        str(r.get("Issue","")),
    ] for r in spacing]

    right_title = "Sharp Bends Without Nearby Vault (<130°, >300 ft)"
    right_headers = ["Conduit ID", "Conduit Vetro ID", "Bend Angle (deg)", "Nearest Vault", "Distance (ft)", "Limit (ft)", "Issue"]
    right_rows = [[
        str(r.get("Conduit ID","")),
        str(r.get("Conduit Vetro ID","")),
        _to_num(r.get("Bend Angle (deg)","")),
        str(r.get("Nearest Vault","")),
        _to_num(r.get("Distance (ft)","")),
        _to_num(r.get("Limit (ft)","")),
        str(r.get("Issue","")),
    ] for r in bends]

    blocks = []
    if left_rows or show_all:
        blocks.append((left_title, left_headers, left_rows))
    if mid_rows or show_all:
        blocks.append((mid_title, mid_headers, mid_rows))
    if right_rows or show_all:
        blocks.append((right_title, right_headers, right_rows))

    # If nothing to show (and not SHOW_ALL), remove the sheet
    if not blocks and not show_all:
        wb.remove(ws)
        return

    # Render selected blocks side-by-side
    GAP = 1
    col = 1
    for title, headers, rows in blocks:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + len(headers) - 1)
        th = ws.cell(row=1, column=col, value=title); th.font = bold; th.alignment = center
        for i, h in enumerate(headers, start=col):
            hc = ws.cell(row=2, column=i, value=h); hc.font = bold; hc.alignment = center
        r = 3
        for row_vals in rows:
            for i, val in enumerate(row_vals, start=col):
                ws.cell(row=r, column=i, value=val)
            r += 1
        col += len(headers) + GAP

    apply_borders(ws)


# --- borders helper ---
def apply_borders(ws):
    """
    Apply borders to a worksheet so that:
      • Header areas: thick outline; thin inner grid.
      • Data areas ("results"): thin outline + thin inner grid.
    Handles custom merged titles, side-by-side blocks, and description banners
    based on the actual content written to the sheet.
    """
    from openpyxl.styles import Border, Side

    thin  = Side(style='thin', color='000000')
    thick = Side(style='thick', color='000000')

    def used_max_row_col():
        max_r = 0
        max_c = 0
        for r in ws.iter_rows():
            for c in r:
                if c.value not in (None, ""):
                    if c.row > max_r: max_r = c.row
                    if c.column > max_c: max_c = c.column
        return max_r, max_c

    def set_cell_border(cell, left=None, right=None, top=None, bottom=None):
        b = cell.border
        cell.border = Border(
            left=left or b.left, right=right or b.right,
            top=top or b.top, bottom=bottom or b.bottom
        )

    def box_outline(min_row, min_col, max_row, max_col, outline_side):
        # draw outline on the rectangle given
        for c in range(min_col, max_col + 1):
            set_cell_border(ws.cell(min_row, c), top=outline_side)
            set_cell_border(ws.cell(max_row, c), bottom=outline_side)
        for r in range(min_row, max_row + 1):
            set_cell_border(ws.cell(r, min_col), left=outline_side)
            set_cell_border(ws.cell(r, max_col), right=outline_side)

    def thin_grid(min_row, min_col, max_row, max_col, outline=False):
        if max_row < min_row or max_col < min_col:
            return
        # apply thin borders to all cells in the rectangle
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(r, c).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        if outline:  # reinforce outline as thin (useful for data boxes)
            box_outline(min_row, min_col, max_row, max_col, thin)

    def last_row_with_data(cols, start_row=1):
        max_r = 0
        max_sheet_row, _ = used_max_row_col()
        for r in range(start_row, max_sheet_row + 1):
            for c in cols:
                v = ws.cell(r, c).value
                if v not in (None, ""):
                    if r > max_r: max_r = r
                    break
        return max_r

    def spans_in_row(row, start_col, end_col):
        """Find contiguous [start,end] column spans where row has any non-empty."""
        spans = []
        in_span = False
        s = None
        for c in range(start_col, end_col + 1):
            v = ws.cell(row, c).value
            if v not in (None, "") and not in_span:
                in_span = True; s = c
            if (v in (None, "")) and in_span:
                spans.append((s, c - 1)); in_span = False
        if in_span:
            spans.append((s, end_col))
        return spans

    def style_header_and_data(header_rows, data_start_row, min_col, max_col):
        # thin grid inside header; thick outline around it
        h_min_row, h_max_row = header_rows
        if h_max_row >= h_min_row:
            thin_grid(h_min_row, min_col, h_max_row, max_col)
            box_outline(h_min_row, min_col, h_max_row, max_col, thick)
        # data grid — thin box and grid
        d_last = last_row_with_data(list(range(min_col, max_col + 1)), data_start_row)
        if d_last and d_last >= data_start_row:
            thin_grid(data_start_row, min_col, d_last, max_col, outline=True)

    max_r, max_c = used_max_row_col()
    if max_r == 0 or max_c == 0:
        return  # nothing to do

    title = (ws.title or "").strip()

    # ───────────── Sheet-specific layouts ─────────────
    if title == 'PON Statistics':
        # Left block: merged banner row 1 + column header row 2; data row 3+ (A:B)
        style_header_and_data((1, 2), 3, 1, 2)
        # NEW Right block: single-row header on row 1; data row 3+ (D:E)
        # ("PON Layers Missing/Present" | "Status")
        style_header_and_data((1, 1), 3, 4, 5)
        return

    if title == 'Drop Issues':
        # Description banner A1:C3; headers row 4; data row 5+
        style_header_and_data((1, 4), 5, 1, 3)
        return

    if title == 'Distribution and NAP Walker':
        # Row 1 headers; data row 2+
        style_header_and_data((1, 1), 2, 1, max_c)
        return

    if title == 'Slack Loop Issues':
        # Summary header rows 1–3 across 4 columns
        style_header_and_data((1, 3), 4, 1, 4)
        # Detail blocks discovered from merged titles on row 5; headers row 6; data row 7+
        row5_merges = [rng for rng in ws.merged_cells.ranges if rng.min_row == 5]
        blocks = []
        if row5_merges:
            for rng in sorted(row5_merges, key=lambda r: r.min_col):
                blocks.append((rng.min_col, rng.max_col))
        else:
            # fallback: contiguous header spans in row 6
            blocks = spans_in_row(6, 1, max_c)

        for (c1, c2) in blocks:
            style_header_and_data((5, 6), 7, c1, c2)
        return

    if title == 'Footage Issues':
        # Two side-by-side blocks with merged titles row 1; headers row 2; data row 3+
        row1_merges = [rng for rng in ws.merged_cells.ranges if rng.min_row == 1]
        if row1_merges:
            for rng in sorted(row1_merges, key=lambda r: r.min_col):
                style_header_and_data((1, 2), 3, rng.min_col, rng.max_col)
        else:
            for (c1, c2) in spans_in_row(2, 1, max_c):
                style_header_and_data((1, 2), 3, c1, c2)
        return

    if title in ('Vaults', 'Conduit'):
        # Multiple side-by-side blocks; merged titles row 1; headers row 2; data row 3+
        row1_merges = [rng for rng in ws.merged_cells.ranges if rng.min_row == 1]
        if row1_merges:
            for rng in sorted(row1_merges, key=lambda r: r.min_col):
                style_header_and_data((1, 2), 3, rng.min_col, rng.max_col)
        else:
            for (c1, c2) in spans_in_row(2, 1, max_c):
                style_header_and_data((1, 2), 3, c1, c2)
        return

    if title == 'NAP Issues':
        # Three blocks across; row 1 has block titles; row 2 has headers; data row 3+
        for (c1, c2) in spans_in_row(2, 1, max_c):
            style_header_and_data((1, 2), 3, c1, c2)
        return

    if title in ('Service Location Issues', 'NID Issues'):
        # Banner row 1 (possibly merged) + header row 2; data row 3+
        spans = spans_in_row(2, 1, max_c)
        if spans:
            for (c1, c2) in spans:
                style_header_and_data((1, 2), 3, c1, c2)
        else:
            style_header_and_data((1, 2), 3, 1, max_c)
        return

    if title in ('Slack Loop Summary', 'GeoJSON Summary'):
        # Metrics grid rows 1–4 (two cols)
        style_header_and_data((1, 4), 5, 1, 2)
        # "Stacked Slack Loops" table: header row begins with 'Coordinate', 'Count'
        header_row = None
        for r in range(1, max_r + 1):
            v1 = ws.cell(r, 1).value
            v2 = ws.cell(r, 2).value
            if (str(v1).strip().lower() == 'coordinate' and
                str(v2).strip().lower() == 'count'):
                header_row = r
                break
        if header_row:
            style_header_and_data((header_row, header_row), header_row + 1, 1, 3)
        return

    if title == 'Power Pole Issues':
        # Merged banner A1:E3; headers row 4; data row 5+
        style_header_and_data((1, 4), 5, 1, max_c)
        return

    # ───────────── Default fallback: treat row 1 as header ─────────────
    style_header_and_data((1, 1), 2, 1, max_c)

def save_workbook(wb, path):
    """
    Save the given Workbook to the specified file path.
    """
    wb.save(path)